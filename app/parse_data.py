from flask import Blueprint, render_template, flash, make_response, current_app

from flask_login import login_required
from flask_wtf import FlaskForm
from wtforms import StringField, IntegerField, SelectField, SubmitField
from wtforms.validators import DataRequired, InputRequired, NumberRange, Length
from flask_wtf.file import FileField, FileAllowed
from werkzeug.utils import secure_filename

from openpyxl import load_workbook
from app.Parser import Parser
import io
import json
import os

bp = Blueprint('parse_data', __name__)


class ParseForm(FlaskForm):
    choices = [
        ('schedule', 'Расписание занятий'),
        ('exams', 'Расписание экзаменов')
    ]

    organization = StringField('Название организации', default='Студенты', validators=[DataRequired(), Length(max=80)])
    header_row = IntegerField('Номер строки заголовков (по умолчанию 0)', default=0, validators=[InputRequired(), NumberRange(min=0)])
    schedule_type = SelectField('Тип расписания', choices=choices, validators=[DataRequired()])
    file = FileField('xlsx файл с расписанием', validators=[DataRequired(), FileAllowed(['xlsx'])])
    submit = SubmitField('Создать JSON файл', render_kw={"onclick": "show_loading_message()"})


@bp.route('/parse_data', methods=['GET', 'POST'])
@login_required
def parse_data():
    form = ParseForm()

    if form.validate_on_submit():
        filename = secure_filename(form.file.data.filename)
        file_path = current_app.config['UPLOAD_DIR_PATH']
        organization = form.organization.data
        header_row = form.header_row.data
        form.file.data.save(file_path + filename)

        json_data = {organization: ''}

        if form.schedule_type.data == 'schedule':
            json_filename = 'schedule_data.json'
        elif form.schedule_type.data == 'exams':
            json_filename = 'exams_data.json'
        else:
            json_filename = 'unknown_data.json'

        try:
            # Предварительная обработка файла (разбиение объединённых ячеек)
            wb = load_workbook(filename=(file_path + filename))

            for sheet in wb.worksheets:
                while sheet.merged_cells.ranges:
                    for merge_range in sheet.merged_cells.ranges:
                        (min_col, min_row, max_col, max_row) = merge_range.bounds
                        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

                        fill_val = sheet.cell(column=min_col, row=min_row).value
                        for col in range(min_col, max_col + 1):
                            for row in range(min_row, max_row + 1):
                                sheet.cell(column=col, row=row, value=fill_val)
            wb.save(file_path + filename)

            # Парсинг обработанного файла с расписанием
            with Parser(file_path + filename) as parser:
                if form.schedule_type.data == 'schedule':
                    data = parser.parse_schedule_from_excel(header_row)
                elif form.schedule_type.data == 'exams':
                    data = parser.parse_exams_from_excel(header_row)
                else:
                    data = []

            json_data = {organization: data}

            flash('JSON файл с расписанием успешно создан')
        except BaseException as e:
            flash(str(e))
            current_app.logger.warning('parse_data: {}'.format(str(e)))
        finally:
            os.remove(os.path.join(file_path, filename))

            with io.StringIO() as json_file:
                json.dump(json_data, json_file, indent=4, sort_keys=True)
                json_file.seek(0)

                response = make_response(json_file.getvalue())
                response.headers['Content-Disposition'] = 'attachment; filename={}'.format(json_filename)
                response.mimetype = 'text/json'

                return response
    return render_template('parse_data.html', title='Парсинг', form=form)
